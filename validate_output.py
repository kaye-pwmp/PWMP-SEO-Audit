#!/usr/bin/env python3
"""
PWMP SEO Audit — output validator.

Runs the "Final Output Checklist" from the workflow against the deliverables
in outputs/<client-slug>/. Exits non-zero if any required check fails so the
audit cannot be declared done with missing or broken outputs.

Usage:
    python scripts/validate_output.py <client-slug>
"""
from __future__ import annotations

import sys
from pathlib import Path

from openpyxl import load_workbook
from pypdf import PdfReader

PROJECT_ROOT = Path(__file__).resolve().parent.parent
OUTPUTS_DIR = PROJECT_ROOT / "outputs"

REQUIRED_TABS = [
    "Navigation",
    "Residential",
    "Commercial",
    "Location Pages",
    "Redirects",
    "Pages to Be Implemented",
    "Blog Transfer",
    "Prompts",
]

REQUIRED_NAV_PAGES = {
    "Home",
    "About Us",
    "Blog",
    "Press",
    "Residential",
    "Commercial",
    "Areas Served",
    "Contact Us",
}


class ValidationError(Exception):
    pass


def _fail(msg: str) -> None:
    print(f"  FAIL  {msg}")


def _ok(msg: str) -> None:
    print(f"  ok    {msg}")


def validate_pdf(pdf_path: Path) -> list[str]:
    errors: list[str] = []
    if not pdf_path.exists():
        errors.append(f"PDF not found at {pdf_path}")
        _fail(errors[-1])
        return errors
    try:
        reader = PdfReader(str(pdf_path))
    except Exception as e:
        errors.append(f"PDF failed to parse: {e}")
        _fail(errors[-1])
        return errors

    pages = len(reader.pages)
    if pages < 12:
        errors.append(f"PDF has only {pages} pages; expected at least 12 (cover + 12 sections).")
        _fail(errors[-1])
    else:
        _ok(f"PDF opens and has {pages} pages")

    # Check footer text on page 2 (cover uses different footer)
    try:
        text = reader.pages[1].extract_text() or ""
        if "Pressure Washing Marketing Pros" not in text or "Confidential" not in text:
            errors.append("PDF footer text missing from body pages.")
            _fail(errors[-1])
        else:
            _ok("PDF footer present on body pages")
    except Exception as e:
        errors.append(f"Could not extract body page text: {e}")
        _fail(errors[-1])

    return errors


def validate_xlsx(xlsx_path: Path) -> list[str]:
    errors: list[str] = []
    if not xlsx_path.exists():
        errors.append(f"XLSX not found at {xlsx_path}")
        _fail(errors[-1])
        return errors
    try:
        wb = load_workbook(xlsx_path)
    except Exception as e:
        errors.append(f"XLSX failed to open: {e}")
        _fail(errors[-1])
        return errors

    if list(wb.sheetnames) != REQUIRED_TABS:
        errors.append(
            f"Sheet order mismatch. Got: {wb.sheetnames}. Expected: {REQUIRED_TABS}"
        )
        _fail(errors[-1])
    else:
        _ok("XLSX has all 8 tabs in correct order")

    nav = wb["Navigation"] if "Navigation" in wb.sheetnames else None
    if nav is not None:
        found_pages = {nav.cell(row=r, column=1).value for r in range(2, nav.max_row + 1)}
        missing = REQUIRED_NAV_PAGES - {p for p in found_pages if p}
        if missing:
            errors.append(f"Navigation tab missing rows: {sorted(missing)}")
            _fail(errors[-1])
        else:
            _ok("Navigation tab has all required rows")

    # Meta description length check across all service/nav tabs
    range_ok = True
    for tab_name in ("Navigation", "Residential", "Commercial", "Location Pages"):
        if tab_name not in wb.sheetnames:
            continue
        ws = wb[tab_name]
        headers = {ws.cell(row=1, column=c).value: c for c in range(1, ws.max_column + 1)}
        meta_col = None
        for key in ("Meta Description", "Meta Description "):
            if key in headers:
                meta_col = headers[key]
                break
        if meta_col is None:
            continue
        for r in range(2, ws.max_row + 1):
            val = ws.cell(row=r, column=meta_col).value
            if not val:
                continue
            length = len(val)
            if not (140 <= length <= 160):
                errors.append(
                    f"{tab_name} row {r}: meta description length {length} out of 140–160 range."
                )
                _fail(errors[-1])
                range_ok = False
    if range_ok:
        _ok("All meta descriptions are in the 140–160 character range")

    # Location Pages should have at least 3 pre-populated rows
    if "Location Pages" in wb.sheetnames:
        ws = wb["Location Pages"]
        filled = sum(1 for r in range(2, 5) if ws.cell(row=r, column=1).value)
        if filled < 3:
            errors.append(
                f"Location Pages tab only has {filled} pre-populated rows; expected 3 top-priority cities."
            )
            _fail(errors[-1])
        else:
            _ok("Location Pages has top 3 cities pre-populated")

    return errors


def main() -> int:
    if len(sys.argv) < 2:
        print("usage: validate_output.py <client-slug>", file=sys.stderr)
        return 2

    slug = sys.argv[1]
    client_dir = OUTPUTS_DIR / slug
    if not client_dir.exists():
        print(f"Client folder not found: {client_dir}", file=sys.stderr)
        return 2

    pdf_path = client_dir / f"{slug}-seo-audit.pdf"
    xlsx_path = client_dir / f"{slug}-website-structure.xlsx"

    print(f"Validating audit outputs for: {slug}")
    print("--- PDF ---")
    pdf_errors = validate_pdf(pdf_path)
    print("--- XLSX ---")
    xlsx_errors = validate_xlsx(xlsx_path)

    total = len(pdf_errors) + len(xlsx_errors)
    print()
    if total == 0:
        print("All checks passed. Outputs are ready to deliver.")
        return 0
    print(f"{total} check(s) failed. Fix the issues above and re-run.")
    return 1


if __name__ == "__main__":
    raise SystemExit(main())
