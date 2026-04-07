#!/usr/bin/env python3
"""
PWMP Website Structure XLSX generator.

Reads a JSON payload with client data and produces an 8-tab .xlsx workbook
matching the PWMP master template. After generation, the teammate uploads
the .xlsx to Google Drive and converts it to a Google Sheet.

Usage:
    python scripts/build_sheet.py --payload path/to/payload.json --out path/to/workbook.xlsx
"""
from __future__ import annotations

import argparse
import json
import re
from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

PWMP_NAVY_HEX = "FF07457C"
PWMP_RED_HEX = "FFBC0300"
LIGHT_GREY_HEX = "FFF2F4F7"

HEADER_FONT = Font(name="Calibri", size=11, bold=True, color="FFFFFFFF")
HEADER_FILL = PatternFill(start_color=PWMP_NAVY_HEX, end_color=PWMP_NAVY_HEX, fill_type="solid")
HEADER_ALIGN = Alignment(horizontal="left", vertical="center", wrap_text=True)
BODY_ALIGN = Alignment(horizontal="left", vertical="top", wrap_text=True)

BLANK_ROWS = 10  # reserved empty rows for manual completion on sparse tabs


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------
def slugify(text: str) -> str:
    text = text.lower().strip()
    text = re.sub(r"[^a-z0-9]+", "-", text)
    return text.strip("-")


def _write_headers(ws, headers: list[str]) -> None:
    for col_idx, h in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=h)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = HEADER_ALIGN
    ws.row_dimensions[1].height = 28
    for col_idx in range(1, len(headers) + 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = 28


def _write_row(ws, row_idx: int, values: list[Any]) -> None:
    for col_idx, v in enumerate(values, start=1):
        cell = ws.cell(row=row_idx, column=col_idx, value=v)
        cell.alignment = BODY_ALIGN


def meta_description(parts: list[str], min_len: int = 140, max_len: int = 160) -> str:
    """
    Build a meta description in the 140-160 char window. Pads with a generic
    CTA if needed so the script never emits an out-of-range meta.
    """
    base = " ".join(p.strip() for p in parts if p).strip()
    base = re.sub(r"\s+", " ", base)
    pad = " Call today for a free quote."
    while len(base) < min_len and len(base) + len(pad) <= max_len:
        base = base + pad if not base.endswith(pad) else base + " Schedule now."
        break
    if len(base) < min_len:
        filler = " Trusted local pros. Call today for a free quote and fast service."
        base = (base + filler)[:max_len]
    if len(base) > max_len:
        base = base[: max_len - 1].rstrip(" ,.;:") + "."
    # Final safety: ensure within window by padding if still short
    if len(base) < min_len:
        base = base + " Call today."
        base = base[:max_len]
    return base


# ---------------------------------------------------------------------------
# Tab 1: Navigation
# ---------------------------------------------------------------------------
def build_navigation(ws, p: dict[str, Any]) -> None:
    headers = ["Page", "URL", "H1", "Alt Text", "Title Tag", "Meta Description"]
    _write_headers(ws, headers)
    business = p.get("business_name", "")
    market = p.get("target_market", "")
    phone = p.get("phone", "")
    btype = p.get("business_type", "pressure washing company")

    rows = [
        [
            "Home",
            "/",
            f"{business} — {btype.title()} in {market}",
            f"{business} {market}",
            f"{business} | {btype.title()} in {market}",
            meta_description(
                [
                    f"{business} offers professional {btype} services in {market}.",
                    f"Get a free quote today — call {phone}.",
                ]
            ),
        ],
        [
            "About Us",
            "/about-us",
            f"About {business}",
            f"About {business} {market}",
            f"About {business} | {btype.title()} in {market}",
            meta_description(
                [
                    f"Learn about {business}, a trusted {btype} serving {market}.",
                    f"Meet our team and call {phone} for a free quote.",
                ]
            ),
        ],
        [
            "Blog",
            "/blog",
            f"{business} Blog — Tips & Guides",
            f"{business} blog {market}",
            f"Blog | {business} — {btype.title()} in {market}",
            meta_description(
                [
                    f"Expert tips and guides from {business}, your trusted {btype} in {market}.",
                    f"Call {phone} for a free estimate.",
                ]
            ),
        ],
        [
            "Press",
            "/press",
            f"{business} In The Press",
            f"{business} press {market}",
            f"Press | {business}",
            meta_description(
                [
                    f"{business} press features and media coverage for our {btype} services in {market}.",
                    f"Call {phone}.",
                ]
            ),
        ],
        ["Residential", "", "[placeholder — see Residential tab]", "", "", ""],
        ["Commercial", "", "[placeholder — see Commercial tab]", "", "", ""],
        [
            "Areas Served",
            "/near-me",
            f"Areas Served by {business}",
            f"{business} service area {market}",
            f"Areas Served | {business} — {btype.title()}",
            meta_description(
                [
                    f"{business} provides {btype} services across {market} and surrounding cities.",
                    f"See our service area and call {phone}.",
                ]
            ),
        ],
        [
            "Contact Us",
            "/contact-us",
            f"Contact {business}",
            f"Contact {business} {market}",
            f"Contact {business} | {btype.title()} in {market}",
            meta_description(
                [
                    f"Contact {business} for {btype} services in {market}.",
                    f"Call {phone} or request a free quote online today.",
                ]
            ),
        ],
    ]
    for i, row in enumerate(rows, start=2):
        _write_row(ws, i, row)


# ---------------------------------------------------------------------------
# Tab 2: Residential
# ---------------------------------------------------------------------------
def _service_rows(services: list[str], p: dict[str, Any]) -> list[list[Any]]:
    business = p.get("business_name", "")
    market = p.get("target_market", "")
    city = p.get("target_city", "")
    state = p.get("target_state", "")
    phone = p.get("phone", "")
    rows: list[list[Any]] = []
    for svc in services or []:
        slug = slugify(svc)
        rows.append(
            [
                svc,
                f"/{slug}/",
                "Use the H1 on the final content brief",
                f"{svc} {city} {state}",
                f"{svc} in {market} | {business}",
                meta_description(
                    [
                        f"Professional {svc.lower()} in {market} by {business}.",
                        f"Call {phone} for a free quote and fast service.",
                    ]
                ),
            ]
        )
    return rows


def build_residential(ws, p: dict[str, Any]) -> None:
    headers = [
        "Residential Service Page",
        "URL",
        "H1",
        "ALT Text",
        "Meta Title",
        "Meta Description",
    ]
    _write_headers(ws, headers)
    for i, row in enumerate(_service_rows(p.get("residential_services", []), p), start=2):
        _write_row(ws, i, row)


# ---------------------------------------------------------------------------
# Tab 3: Commercial
# ---------------------------------------------------------------------------
def build_commercial(ws, p: dict[str, Any]) -> None:
    headers = ["Service", "URL", "H1", "Meta Title", "Meta Description", "Alt Text"]
    _write_headers(ws, headers)
    business = p.get("business_name", "")
    market = p.get("target_market", "")
    city = p.get("target_city", "")
    state = p.get("target_state", "")
    phone = p.get("phone", "")
    row_idx = 2
    for svc in p.get("commercial_services", []) or []:
        slug = slugify(svc)
        _write_row(
            ws,
            row_idx,
            [
                svc,
                f"/{slug}/",
                "Use the H1 on the final content brief",
                f"{svc} in {market} | {business}",
                meta_description(
                    [
                        f"Reliable {svc.lower()} in {market} by {business}.",
                        f"Call {phone} for commercial pricing and scheduling.",
                    ]
                ),
                f"{svc} {city} {state}",
            ],
        )
        row_idx += 1


# ---------------------------------------------------------------------------
# Tab 4: Location Pages
# ---------------------------------------------------------------------------
def build_locations(ws, p: dict[str, Any]) -> None:
    headers = ["Location Name", "URL", "H1", "Meta Title", "Meta Description", "Alt Text"]
    _write_headers(ws, headers)
    business = p.get("business_name", "")
    phone = p.get("phone", "")
    btype = p.get("business_type", "pressure washing company")
    primary_service = None
    if p.get("residential_services"):
        primary_service = p["residential_services"][0]
    elif p.get("commercial_services"):
        primary_service = p["commercial_services"][0]
    primary_service = primary_service or "Pressure Washing"
    service_slug = slugify(primary_service)

    row_idx = 2
    for loc in p.get("top_3_cities", []) or []:
        city = loc.get("city", "")
        state = loc.get("state", "")
        city_slug = slugify(city)
        state_slug = slugify(state)
        _write_row(
            ws,
            row_idx,
            [
                f"{city}, {state}",
                f"/near-me/{service_slug}-{city_slug}-{state_slug}/",
                f"{primary_service} in {city}, {state}",
                f"{primary_service} in {city}, {state} | {business}",
                meta_description(
                    [
                        f"{business} provides {primary_service.lower()} in {city}, {state}.",
                        f"Locally trusted {btype}. Call {phone} for a free quote.",
                    ]
                ),
                f"{primary_service} {city} {state}",
            ],
        )
        row_idx += 1

    for _ in range(BLANK_ROWS):
        _write_row(ws, row_idx, ["", "", "", "", "", ""])
        row_idx += 1
    _write_row(ws, row_idx, ["Extra Locations", "", "", "", "", ""])


# ---------------------------------------------------------------------------
# Tab 5: Redirects
# ---------------------------------------------------------------------------
def build_redirects(ws, p: dict[str, Any]) -> None:
    headers = [
        "Old URL (Put the URL that Needs to Be Redirected)",
        "New URL (Put URL That Is Being Redirected to)",
    ]
    _write_headers(ws, headers)
    ws.column_dimensions["A"].width = 55
    ws.column_dimensions["B"].width = 55
    row_idx = 2
    for r in p.get("recommended_redirects", []) or []:
        _write_row(ws, row_idx, [r.get("old", ""), r.get("new", "")])
        row_idx += 1
    for _ in range(BLANK_ROWS):
        _write_row(ws, row_idx, ["", ""])
        row_idx += 1


# ---------------------------------------------------------------------------
# Tab 6: Pages to Be Implemented
# ---------------------------------------------------------------------------
def build_pages_to_implement(ws, p: dict[str, Any]) -> None:
    headers = ["Current Page URL", "URL on New Website"]
    _write_headers(ws, headers)
    ws.column_dimensions["A"].width = 55
    ws.column_dimensions["B"].width = 55
    row_idx = 2
    for url in p.get("existing_pages", []) or []:
        _write_row(ws, row_idx, [url, ""])
        row_idx += 1
    for _ in range(BLANK_ROWS):
        _write_row(ws, row_idx, ["", ""])
        row_idx += 1


# ---------------------------------------------------------------------------
# Tab 7: Blog Transfer (with TRUE/FALSE data validation)
# ---------------------------------------------------------------------------
def build_blog_transfer(ws, p: dict[str, Any]) -> None:
    headers = ["Blog URL", "Will Transfer?", "Notes"]
    _write_headers(ws, headers)
    ws.column_dimensions["A"].width = 55
    ws.column_dimensions["B"].width = 16
    ws.column_dimensions["C"].width = 40

    dv = DataValidation(type="list", formula1='"TRUE,FALSE"', allow_blank=True)
    dv.error = "Please choose TRUE or FALSE"
    dv.errorTitle = "Invalid entry"
    ws.add_data_validation(dv)

    row_idx = 2
    posts = p.get("existing_blog_posts", []) or []
    for post in posts:
        _write_row(ws, row_idx, [post, "", ""])
        row_idx += 1
    for _ in range(BLANK_ROWS):
        _write_row(ws, row_idx, ["", "", ""])
        row_idx += 1
    dv.add(f"B2:B{row_idx - 1}")


# ---------------------------------------------------------------------------
# Tab 8: Prompts
# ---------------------------------------------------------------------------
HOMEPAGE_META_PROMPT = (
    "Write a compelling 140–160 character homepage meta description for [business name], "
    "a [type of business] serving [location]. Include the primary service, the city, "
    "and a clear call to action with the phone number [insert client's phone number]. "
    "Avoid generic filler and keep it conversion-focused."
)

CONTACT_PAGE_PROMPT = (
    "Write the Contact Us page copy for [business name], a [type of business] in [location]. "
    "Include: a one-paragraph intro, the phone number [insert client's phone number] as a "
    "clickable CTA, business hours, service areas, and a reassurance line about free quotes "
    "and fast response. Tone should be warm, professional, and local."
)

META_BATCH_PROMPT = (
    "You are writing meta titles and meta descriptions for [business name], a [type of business] "
    "serving [location]. For each service or location page I provide, generate:\n"
    "1) A meta title formatted as: <Page Focus> in <Location> | [business name]  (under 60 chars)\n"
    "2) A meta description that is UNIQUE, 140–160 characters, includes the primary keyword, the city, "
    "the phone number [insert client's phone number], and a specific call to action.\n"
    "Never repeat phrasing across pages. Never exceed 160 characters on descriptions."
)


def _replace_placeholders(text: str, p: dict[str, Any]) -> str:
    return (
        text.replace("[business name]", p.get("business_name", "[business name]"))
        .replace("[type of business]", p.get("business_type", "[type of business]"))
        .replace("[location]", p.get("target_market", "[location]"))
        .replace("[insert client's phone number]", p.get("phone", "[insert client's phone number]"))
    )


def build_prompts(ws, p: dict[str, Any]) -> None:
    headers = ["Prompt Type", "Prompt"]
    _write_headers(ws, headers)
    ws.column_dimensions["A"].width = 32
    ws.column_dimensions["B"].width = 90

    rows = [
        ("Homepage meta description prompt", _replace_placeholders(HOMEPAGE_META_PROMPT, p)),
        ("Contact Us page prompt", _replace_placeholders(CONTACT_PAGE_PROMPT, p)),
        ("Meta Titles and Descriptions batch prompt", _replace_placeholders(META_BATCH_PROMPT, p)),
    ]
    for i, (label, text) in enumerate(rows, start=2):
        ws.cell(row=i, column=1, value=label).alignment = BODY_ALIGN
        cell = ws.cell(row=i, column=2, value=text)
        cell.alignment = BODY_ALIGN
        ws.row_dimensions[i].height = 110


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------
TAB_ORDER = [
    ("Navigation", build_navigation),
    ("Residential", build_residential),
    ("Commercial", build_commercial),
    ("Location Pages", build_locations),
    ("Redirects", build_redirects),
    ("Pages to Be Implemented", build_pages_to_implement),
    ("Blog Transfer", build_blog_transfer),
    ("Prompts", build_prompts),
]


def main() -> int:
    ap = argparse.ArgumentParser(description="Generate PWMP website structure XLSX.")
    ap.add_argument("--payload", required=True)
    ap.add_argument("--out", required=True)
    args = ap.parse_args()

    payload_path = Path(args.payload)
    out_path = Path(args.out)
    out_path.parent.mkdir(parents=True, exist_ok=True)

    with payload_path.open("r", encoding="utf-8") as f:
        payload = json.load(f)

    wb = Workbook()
    wb.remove(wb.active)  # we add all tabs explicitly

    for name, builder in TAB_ORDER:
        ws = wb.create_sheet(title=name)
        builder(ws, payload)

    wb.save(out_path)
    print(f"[build_sheet] Wrote {out_path}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
